#!/usr/bin/env python3
"""Build the downloadable bundle sold via Gumroad.

Output dir: data/product_bundle/
  - hsnlookup-india-2026.xlsx   — Excel with formulas for landed cost
  - hsnlookup-india-2026.csv    — UTF-8 CSV of the raw rate data
  - hsnlookup-india-2026.json   — JSON equivalent, same schema as /api/hsn.json
  - README.txt                  — licence, changelog, how to use, support inbox

Packaged later via `cd data/product_bundle && zip -r ../product_bundle.zip .`
"""
import json
import os
import csv
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
OUT = ROOT / "data" / "product_bundle"
OUT.mkdir(parents=True, exist_ok=True)

seed = json.loads((ROOT / "data/seed/hsn_expanded.json").read_text())
codes = seed["codes"]
codes.sort(key=lambda c: c["hsn"])

# CSV
csv_path = OUT / "hsnlookup-india-2026.csv"
with csv_path.open("w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["hsn", "description", "chapter", "bcd_pct", "sws_pct_of_bcd", "igst_pct", "cess_pct", "igst_verification", "bcd_verification", "as_of"])
    for c in codes:
        v = c.get("verification") or {}
        w.writerow([
            c["hsn"], c["description"], c["chapter"],
            c["bcd"], c["sws"], c["igst"], c.get("cess", 0),
            v.get("igst", ""), v.get("bcd", ""), v.get("asOf", ""),
        ])

# JSON
json_path = OUT / "hsnlookup-india-2026.json"
json_path.write_text(json.dumps({
    "generatedAt": date.today().isoformat(),
    "source": "hsnlookup.in",
    "license": "CC-BY-4.0 — attribute hsnlookup.in with a visible link in any redistributed form",
    "count": len(codes),
    "schema": {
        "hsn": "string — India ITC(HS) 8-digit tariff item",
        "description": "string",
        "chapter": "number 1..97",
        "bcd": "Basic Customs Duty percent",
        "sws": "Social Welfare Surcharge percent of BCD",
        "igst": "Integrated GST percent",
        "cess": "Compensation Cess percent",
    },
    "data": [
        {k: c[k] for k in ("hsn", "description", "chapter", "bcd", "sws", "igst", "cess") if k in c}
        for c in codes
    ],
}, indent=2))

# Excel — with worked-formula columns so buyer can plug in a CIF and see landed cost
wb = Workbook()
ws = wb.active
ws.title = "HSN Rates"

headers = ["HSN", "Description", "Chapter", "BCD %", "SWS % of BCD", "IGST %", "Cess %", "CIF (₹)", "AV (₹)", "BCD (₹)", "SWS (₹)", "Cess (₹)", "IGST (₹)", "Total Duty (₹)", "Landed Cost (₹)", "Effective %"]
ws.append(headers)

header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Data rows, with CIF defaulting to 100000 so formulas produce meaningful numbers out of the box
for i, c in enumerate(codes, start=2):
    ws.cell(row=i, column=1, value=c["hsn"])
    ws.cell(row=i, column=2, value=c["description"])
    ws.cell(row=i, column=3, value=c["chapter"])
    ws.cell(row=i, column=4, value=c["bcd"])
    ws.cell(row=i, column=5, value=c["sws"])
    ws.cell(row=i, column=6, value=c["igst"])
    ws.cell(row=i, column=7, value=c.get("cess", 0))
    ws.cell(row=i, column=8, value=100000)  # CIF default
    # AV = CIF + 1% landing
    ws.cell(row=i, column=9, value=f"=H{i}*1.01")
    # BCD = AV * bcd%
    ws.cell(row=i, column=10, value=f"=I{i}*D{i}/100")
    # SWS = BCD * sws%
    ws.cell(row=i, column=11, value=f"=J{i}*E{i}/100")
    # Cess = (AV+BCD) * cess%
    ws.cell(row=i, column=12, value=f"=(I{i}+J{i})*G{i}/100")
    # IGST = (AV+BCD+SWS+Cess) * igst%
    ws.cell(row=i, column=13, value=f"=(I{i}+J{i}+K{i}+L{i})*F{i}/100")
    # Total duty
    ws.cell(row=i, column=14, value=f"=J{i}+K{i}+L{i}+M{i}")
    # Landed
    ws.cell(row=i, column=15, value=f"=I{i}+N{i}")
    # Effective %
    ws.cell(row=i, column=16, value=f"=N{i}/I{i}*100")

# Column widths
widths = {1: 12, 2: 60, 3: 8, 4: 9, 5: 11, 6: 9, 7: 9, 8: 14, 9: 14, 10: 12, 11: 11, 12: 11, 13: 12, 14: 14, 15: 14, 16: 11}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w
ws.freeze_panes = "A2"

# Readme sheet
ws2 = wb.create_sheet("README")
readme = [
    ["hsnlookup.in — India HSN + Customs Duty Database 2026"],
    [""],
    ["What's in this file"],
    [f"{len(codes):,} India 8-digit ITC(HS) tariff items with BCD, SWS, IGST and cess rates, plus live landed-cost formulas."],
    [""],
    ["How to use"],
    ["1. Change the CIF value in column H for any row. The AV, BCD, SWS, Cess, IGST, Total Duty, Landed Cost and Effective % columns update automatically."],
    ["2. Columns D-G (BCD/SWS/IGST/Cess percentages) are the source rates. Edit them to model alternate scenarios (FTA concessional rates, anti-dumping overlays, etc)."],
    ["3. Column P (Effective %) = (Total Duty / AV) × 100. The all-in duty rate after the four components stack together."],
    [""],
    ["Data sources"],
    ["• IGST: CBIC GST rate schedule (notifications under the IGST Act, 2017), compiled via Cleartax's public HSN lookup API."],
    ["• BCD: hand-reviewed for 36 high-interest codes; chapter-default First Schedule statutory rate for the rest (verify against current CBIC notification before filing Bill of Entry)."],
    ["• SWS: 10% of BCD, Finance Act 2018 Section 110."],
    ["• Cess: Compensation Cess under GST Compensation Act 2017 for specific goods."],
    [""],
    ["Licence"],
    ["CC-BY-4.0. You may use this in commercial tools, ERPs, dashboards. You must attribute hsnlookup.in with a visible link back."],
    [""],
    ["Support"],
    ["hello@hsnlookup.in — rate corrections, bulk-licence enquiries, custom formats."],
    [""],
    ["Generated"],
    [date.today().isoformat() + " • hsnlookup.in"],
]
for i, row in enumerate(readme, start=1):
    ws2.cell(row=i, column=1, value=row[0])
ws2.column_dimensions["A"].width = 120
ws2["A1"].font = Font(bold=True, size=16, color="1D4ED8")

wb.save(OUT / "hsnlookup-india-2026.xlsx")

# README.txt as plain text too
readme_text = "\n".join(r[0] for r in readme)
(OUT / "README.txt").write_text(readme_text + "\n")

print(f"Wrote bundle to {OUT}/")
print(f"  - hsnlookup-india-2026.xlsx ({(OUT/'hsnlookup-india-2026.xlsx').stat().st_size // 1024} KB)")
print(f"  - hsnlookup-india-2026.csv ({(OUT/'hsnlookup-india-2026.csv').stat().st_size // 1024} KB)")
print(f"  - hsnlookup-india-2026.json ({(OUT/'hsnlookup-india-2026.json').stat().st_size // 1024} KB)")
print(f"  - README.txt")
